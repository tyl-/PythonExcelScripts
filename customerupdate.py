#!/usr/bin/python

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os

# Global variables
months = ["01 - January", "02 - February", "03 - March", "04 - April", "05 - May", "06 - June", "07 - July", "08 - August", "09 - September", "10 - October", "11 - November", "12 - December"]
# Change year as needed
year = 2017
currentPath = os.path.abspath(os.curdir).replace('\\', '/') + '/'
dailyDir = ""
customerDir = ""

# Assumes structure is the same as genenerated by generate.py
def updateDirPath():
    global dailyDir, customerDir
    print "Updating file paths..."
    if os.path.exists(currentPath + "Daily Report " + str(year) + "/"):
        dailyDir = currentPath + "Daily Report " + str(year) + "/"
        customerDir = currentPath + "Specific Report " + str(year) + "/NumberOfCustomer " + str(year) + "/"
    else:
        dailyDir = os.path.abspath(os.path.join(os.pardir,os.pardir)).replace('\\', '/') + '/'
        customerDir = dailyDir
        dailyDir += "Daily Report " + str(year) + "/"
        customerDir += "Specific Report " + str(year) + "/NumberOfCustomer " + str(year) + "/"
    print " Daily files assumed to be at \"" + dailyDir + "\""
    print " Number of customer files assumed to be at \"" + customerDir + "\""

# Assumes excel files are based on the templates
def updateExcel():
    for file in os.listdir(customerDir):
        if file.endswith(".xlsx"):
            print file
            wb = load_workbook(filename = customerDir + file)
            ws = wb.active
            for i in range(3, 53):
                if(ws['B' + str(i)].value != None) and (ws['C' +str(i)].value != None):
                    print "files to read", ws['B'+str(i)].value, ws['C'+str(i)].value
                    mS, dS, yS = ws['B' + str(i)].value.split("-")
                    mE, dE, yE = ws['C' + str(i)].value.split("-")
                    #print mS + dS + yS + " " + mE + dE + yE
                    nMS = int(mS)
                    nDS = int(dS)
                    nYS = int(yS)
                    nME = int(mE)
                    nDE = int(dE)
                    nYE = int(yE)
                    #print nMS, nDS, nYS, nME, nDE, nYE
                    numCustomer = 0
                    while(True):
                        if(nYS > nYE):
                            break;
                        elif(nYS == nYE) and (nMS > nME):
                            break;
                        elif(nMS == nME) and (nDS > nDE):
                            break;

                        if(nMS < 10):
                            sMS = "0" + str(nMS)
                        else:
                            sMS = str(nMS)
                        if(nDS < 10):
                            sDS = "0" + str(nDS)
                        else:
                            sDS = str(nDS)

                        sF = sMS + "-" + sDS + "-" + str(nYS) + ".xlsx"
                        print sF

                        if(nMS == 1) or (nMS == 3) or (nMS == 5) or (nMS ==7) or (nMS == 8) or (nMS == 10) or (nMS == 12):
                            if(nDS < 31):
                                nDS += 1
                            elif(nDS == 31) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        elif(nMS == 4) or (nMS == 6) or (nMS == 9) or (nMS == 11):
                            if(nDS < 30):
                                nDS += 1
                            elif(nDS == 30) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        elif(nMS == 2) and (nYS % 4) == 0:
                            if(nDS < 29):
                                nDS += 1
                            elif(nDS == 29) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        elif(nMS == 2) and (nYS % 4) != 0:
                            if(nDS < 28):
                                nDS += 1
                            elif(nDS == 28) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        else:
                            break;

                        wb2 = load_workbook(filename = dailyDir + months[int(sMS) - 1] + "/" + sF, data_only=True)
                        ws2 = wb2.active
                        print "# of customer ", ws2['D19'].value
                        numCustomer += ws2['D19'].value

                    print "Total # customer", numCustomer
                    ws['D'+str(i)].value = numCustomer
                    wb.save(customerDir + file)
            for i in range(3, 53):
                if(ws['F' + str(i)].value != None) and (ws['G' +str(i)].value != None):
                    print "files to read", ws['B'+str(i)].value, ws['C'+str(i)].value
                    mS, dS, yS = ws['F' + str(i)].value.split("-")
                    mE, dE, yE = ws['G' + str(i)].value.split("-")
                    #print mS + dS + yS + " " + mE + dE + yE
                    nMS = int(mS)
                    nDS = int(dS)
                    nYS = int(yS)
                    nME = int(mE)
                    nDE = int(dE)
                    nYE = int(yE)
                    #print nMS, nDS, nYS, nME, nDE, nYE
                    numCustomer = 0
                    while(True):
                        if(nYS > nYE):
                            break;
                        elif(nYS == nYE) and (nMS > nME):
                            break;
                        elif(nMS == nME) and (nDS > nDE):
                            break;

                        if(nMS < 10):
                            sMS = "0" + str(nMS)
                        else:
                            sMS = str(nMS)
                        if(nDS < 10):
                            sDS = "0" + str(nDS)
                        else:
                            sDS = str(nDS)

                        sF = sMS + "-" + sDS + "-" + str(nYS) + ".xlsx"
                        print sF

                        if(nMS == 1) or (nMS == 3) or (nMS == 5) or (nMS ==7) or (nMS == 8) or (nMS == 10) or (nMS == 12):
                            if(nDS < 31):
                                nDS += 1
                            elif(nDS == 31) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        elif(nMS == 4) or (nMS == 6) or (nMS == 9) or (nMS == 11):
                            if(nDS < 30):
                                nDS += 1
                            elif(nDS == 30) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        elif(nMS == 2) and (nYS % 4) == 0:
                            if(nDS < 29):
                                nDS += 1
                            elif(nDS == 29) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        elif(nMS == 2) and (nYS % 4) != 0:
                            if(nDS < 28):
                                nDS += 1
                            elif(nDS == 28) and (nMS < 12):
                                nMS += 1
                                nDS = 1
                            elif(nMS == 12) and (nYS < nYE):
                                nYS += 1
                                nMS = 1
                                nDS = 1
                            else:
                                break;
                        else:
                            break;

                        wb2 = load_workbook(filename = dailyDir + months[int(sMS) - 1] + "/" + sF, data_only=True)
                        ws2 = wb2.active
                        print "# of customer ", ws2['D19'].value
                        numCustomer += ws2['D19'].value

                    print "Total # customer", numCustomer
                    ws['H'+str(i)].value = numCustomer
                    wb.save(customerDir + file)

                
        


def fixBorders():
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),bottom=Side(style="thin"))
    for file in os.listdir(customerDir):
        if file.endswith(".xlsx"):
            print file
            wb = load_workbook(filename = customerDir + file)
            ws = wb.active
            bp = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']
            for i in bp:
                ws[i].border = thin_border
            wb.save(customerDir + file)
    print("Fixed border styles")
    
def main():
    print currentPath
    updateDirPath()
    updateExcel()
    fixBorders()

if __name__ == "__main__":
    main()
