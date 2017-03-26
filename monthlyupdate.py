#!/usr/bin/python

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os

# Global variables
months = ["01 - January", "02 - February", "03 - March", "04 - April", "05 - May", "06 - June", "07 - July", "08 - August", "09 - September", "10 - October", "11 - November", "12 - December"]
# Change year as needed
year = 2017
currentPath = os.path.abspath(os.curdir).replace('\\', '/') + '/'
dailyDir = ""
monthlyDir = ""

# Assumes structure is the same as genenerated by generate.py
def updateDirPath():
    global dailyDir, monthlyDir
    print "Updating file paths..."
    if os.path.exists(currentPath + "Daily Report " + str(year) + "/"):
        dailyDir = currentPath + "Daily Report " + str(year) + "/"
        monthlyDir = currentPath + "Monthly Report " + str(year) + "/"
    else:
        dailyDir = os.path.abspath(os.path.join(os.pardir,os.curdir)).replace('\\', '/') + '/'
        monthlyDir = dailyDir
        dailyDir += "Daily Report " + str(year) + "/"
        monthlyDir += "Monthly Report " + str(year) + "/"
    print " Daily files assumed to be at \"" + dailyDir + "\""
    print " Monthly files assumed to be at \"" + monthlyDir + "\""

# Assumes excel files are based on the templates
def updateExcel():
    global months
    for month in months:
        print "Month is", month
        wb = load_workbook(filename = monthlyDir + month + ' 2017.xlsx')
        ws = wb.active
        for file in os.listdir(dailyDir + month + "/"):
            m, d, y = file.split("-")
            dN = int(d) + 4
            i = str(dN)
            if file.endswith(".xlsx"):
                m, d, y = file.split("-")
                dN = int(d) + 4
                i = str(dN)
                print "Loading excel file", file
                wb2 = load_workbook(filename = dailyDir + month + "/" + file, data_only=True)
                ws2 = wb.active
                ws['B' + i].value = ws2['B5'].value
                ws['C' + i].value = ws2['C5'].value
                #ws['D' + i].value = ws2['D5'].value
                ws['E' + i].value = ws2['B14'].value
                ws['F' + i].value = ws2['C14'].value
                #ws['G' + i].value = ws2['D14'].value
                ws['H' + i].value = ws2['B6'].value
                ws['I' + i].value = ws2['C6'].value
                #ws['J' + i].value = ws2['D6'].value
                ws['K' + i].value = ws2['B15'].value
                ws['L' + i].value = ws2['C15'].value
                #ws['M' + i].value = ws2['D15'].value
                ws['N' + i].value = ws2['B19'].value
                ws['O' + i].value = ws2['C19'].value
                #ws['P' + i].value = ws2['D19'].value
                wb.save(monthlyDir + month + " 2017.xlsx")
                print("Saved changes to", monthlyDir + month + " 2017.xlsx")

def fixBorders():
    medium_border = Border(left=Side(style='medium'), right=Side(style="medium"), top=Side(style="medium"),bottom=Side(style="medium"))
    for month in months:
        wb = load_workbook(filename = monthlyDir + month + ' 2017.xlsx')
        ws = wb.active
        bp = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
        for i in bp[1:]:
            ws[i + '3'].border = medium_border
        if(month[:2] == "01" or month[:2] == "03" or month[:2] == "05" or month[:2] == "07" or month[:2] == "08" or month[:2] == "10" or month[:2] == "12"):
            for i in bp[0:len(bp)-3]:
                ws[i + '37'].border = medium_border
        elif(month[:2] == "04" or month[:2] == "06" or month[:2] == "09" or month[:2] == "11"):
            for i in bp[0:len(bp)-3]:
                ws[i + '36'].border = medium_border
        elif(month[:2] == "02"):
            if(year % 4 == 0):
                for i in bp[0:len(bp)-3]:
                    ws[i + '35'].border = medium_border
            else:
                for i in bp[0:len(bp)-3]:
                    ws[i + '34'].border = medium_border
        wb.save(monthlyDir + month + " 2017.xlsx")
    print("Fixed border styles")
    
def main():
    print currentPath
    updateDirPath()
    updateExcel()
    fixBorders()

if __name__ == "__main__":
    main()
            
